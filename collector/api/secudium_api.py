"""
SECUDIUM API Client
SK Shielders Security Intelligence Platform Collector

This module handles:
1. Authentication with SECUDIUM platform
2. Fetching daily blacklist IP reports
3. Downloading Excel files via BROWSER AUTOMATION (API endpoint blocks downloads)
4. Parsing Excel files
5. Storing data in PostgreSQL database

Author: Blacklist System Team
Date: 2025-10-19
Version: 1.1.0 - Browser automation for downloads

IMPORTANT: The /file/SECINFO/download API endpoint returns
"관리자에게 문의해 주세요" error even with valid authentication.
Browser automation is used as workaround.
"""

import requests
import pandas as pd
import io
from typing import List, Dict, Optional, Tuple
from datetime import datetime, timedelta
import structlog
from openpyxl import load_workbook
import os
import shutil
import tempfile

logger = structlog.get_logger()


class SecudiumAPIClient:
    """
    Client for SECUDIUM (SK Shielders) security intelligence platform.

    Similar to REGTECH collector but with SECUDIUM-specific authentication
    and data format handling.

    API Endpoints:
        - POST /isap-api/loginProcess          # Authentication
        - GET  /isap-api/secinfo/list/black_ip # List reports
        - GET  /isap-api/secinfo/view/black_ip/{id} # Report details
        - GET  /isap-api/file/SECINFO/download # Download Excel

    Usage:
        >>> client = SecudiumAPIClient(username="user", password="pass")
        >>> client.authenticate()
        >>> reports = client.get_latest_reports(limit=5)
        >>> for report in reports:
        ...     data = client.process_report(report['id'])
    """

    BASE_URL = "https://secudium.skinfosec.co.kr/isap-api"
    TIMEOUT = 30  # seconds
    MAX_RETRIES = 3

    def __init__(self, username: str, password: str):
        """
        Initialize SECUDIUM API client.

        Args:
            username: SECUDIUM platform username
            password: SECUDIUM platform password
        """
        self.username = username
        self.password = password
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'ko-KR,ko;q=0.9',
            'Referer': 'https://secudium.skinfosec.co.kr/',
            'Origin': 'https://secudium.skinfosec.co.kr'
        })
        self._authenticated = False
        self._token = None  # Session token from login response

        logger.info("secudium_client_initialized", username=username)

    def authenticate(self) -> bool:
        """
        Authenticate with SECUDIUM platform.

        The API uses session-based authentication. Successful login
        sets a session cookie that's used for subsequent requests.

        Returns:
            bool: True if authentication successful

        Raises:
            requests.HTTPError: If authentication fails
            requests.Timeout: If request times out
        """
        # Skip if already authenticated
        if self._authenticated:
            logger.info("secudium_already_authenticated", username=self.username)
            return True

        try:
            logger.info("secudium_auth_attempt", username=self.username)

            # SECUDIUM uses form-urlencoded with specific field names
            payload = {
                "login_name": self.username,  # NOT 'username'!
                "password": self.password,
                "lang": "ko",                 # Language setting
                "is_expire": "Y",             # Force logout existing sessions
                "is_otp": "N"                 # No OTP required
            }

            response = self.session.post(
                f"{self.BASE_URL}/loginProcess",
                data=payload,
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                timeout=self.TIMEOUT
            )

            # Debug: Log response headers to see if cookies are being set
            logger.info("secudium_auth_response_headers",
                       status_code=response.status_code,
                       has_set_cookie=('Set-Cookie' in response.headers),
                       set_cookie_value=response.headers.get('Set-Cookie', 'None'),
                       all_headers=list(response.headers.keys()))

            # SECUDIUM returns 200 with JSON (not 204!)
            if response.status_code == 200:
                try:
                    data = response.json()
                    resp_data = data.get("response", {})

                    # Check for errors
                    if resp_data.get("error"):
                        # Handle duplicate login
                        if resp_data.get("code") == "already.login":
                            logger.warning("secudium_duplicate_login",
                                         username=self.username,
                                         message="Existing session found, forcing logout...")

                            # Retry with force logout
                            payload["is_expire"] = "Y"
                            response = self.session.post(
                                f"{self.BASE_URL}/loginProcess",
                                data=payload,
                                headers={"Content-Type": "application/x-www-form-urlencoded"},
                                timeout=self.TIMEOUT
                            )

                            if response.status_code == 200:
                                data2 = response.json()
                                resp_data2 = data2.get("response", {})

                                if not resp_data2.get("error"):
                                    self._authenticated = True
                                    self._token = resp_data2.get("token")

                                    # Debug: Log cookies after forced authentication
                                    cookies_dict = dict(self.session.cookies)
                                    logger.info("secudium_auth_success_forced",
                                              username=self.username,
                                              token_prefix=self._token[:30] if self._token else None,
                                              num_cookies=len(cookies_dict),
                                              cookie_names=list(cookies_dict.keys()))
                                    return True

                        # Other errors
                        error_msg = resp_data.get("message", "Unknown error")
                        logger.error("secudium_auth_failed",
                                   username=self.username,
                                   error_code=resp_data.get("code"),
                                   error_message=error_msg)
                        raise ValueError(f"Authentication failed: {error_msg}")

                    # Success - no error
                    else:
                        self._authenticated = True
                        self._token = resp_data.get("token")

                        # Debug: Log cookies after authentication
                        cookies_dict = dict(self.session.cookies)
                        logger.info("secudium_auth_success",
                                  username=self.username,
                                  token_prefix=self._token[:30] if self._token else None,
                                  num_cookies=len(cookies_dict),
                                  cookie_names=list(cookies_dict.keys()))

                        return True

                except (ValueError, KeyError) as e:
                    logger.error("secudium_auth_response_parse_error",
                               username=self.username,
                               error=str(e))
                    raise
            else:
                response.raise_for_status()

        except requests.RequestException as e:
            logger.error("secudium_auth_failed",
                        username=self.username,
                        error=str(e),
                        status_code=getattr(e.response, 'status_code', None))
            raise

    def get_latest_reports(self, limit: int = 10, days_back: int = 7) -> List[Dict]:
        """
        Fetch latest blacklist reports from SECUDIUM.

        Reports are returned in reverse chronological order (newest first).
        Each report represents a daily blacklist update from SK-SOC.

        Args:
            limit: Maximum number of reports to fetch
            days_back: Only fetch reports from last N days

        Returns:
            List of report metadata dictionaries with fields:
                - id: Report ID
                - title: Report title (includes date)
                - author: Report author
                - date: Publication date
                - file_uuid: File UUID for download
                - filename: Original Excel filename

        Raises:
            RuntimeError: If not authenticated
        """
        if not self._authenticated:
            raise RuntimeError("Must authenticate before fetching reports")

        try:
            logger.info("secudium_fetch_reports", limit=limit)

            # Add token to headers if available
            headers = {}
            if self._token:
                headers["X-Auth-Token"] = self._token

            response = self.session.get(
                f"{self.BASE_URL}/secinfo/list/black_ip",
                headers=headers,
                timeout=self.TIMEOUT
            )
            response.raise_for_status()

            data = response.json()
            all_reports = data.get("rows", [])

            # Filter by date and limit
            cutoff_date = datetime.now() - timedelta(days=days_back)
            filtered_reports = []

            for report in all_reports[:limit]:
                report_data = report.get("data", [])
                if len(report_data) >= 5:
                    # Parse report metadata from data array
                    # Format: [empty, empty, title, author, date, download_btn, ...]
                    report_date_str = report_data[4]  # "2025-10-17 01:26:06"

                    try:
                        report_date = datetime.strptime(
                            report_date_str,
                            "%Y-%m-%d %H:%M:%S"
                        )

                        if report_date >= cutoff_date:
                            filtered_reports.append({
                                'id': report['id'],
                                'title': report_data[2],
                                'author': report_data[3],
                                'date': report_date,
                                'raw': report
                            })
                    except ValueError as e:
                        logger.warning("secudium_date_parse_error",
                                     report_id=report['id'],
                                     date_str=report_date_str,
                                     error=str(e))
                        continue

            logger.info("secudium_reports_fetched",
                       total=len(all_reports),
                       filtered=len(filtered_reports),
                       days_back=days_back)

            return filtered_reports

        except requests.RequestException as e:
            logger.error("secudium_fetch_failed", error=str(e))
            raise

    def get_report_details(self, report_id: int) -> Dict:
        """
        Get detailed information for a specific report.

        This includes the file UUID and filename needed for downloading
        the Excel file containing blacklist IPs.

        Args:
            report_id: SECUDIUM report ID

        Returns:
            Report details dictionary with fields:
                - seq: Report sequence number
                - title: Report title
                - content: HTML content
                - sefileString255: File UUID (for download)
                - fileString255: Original filename
                - regDate: Registration timestamp
                - ip: Publisher IP address (NOT blacklist data)

        Raises:
            RuntimeError: If not authenticated
        """
        if not self._authenticated:
            raise RuntimeError("Must authenticate before fetching report details")

        try:
            logger.info("secudium_fetch_details", report_id=report_id)

            # Add token to headers if available
            headers = {}
            if self._token:
                headers["X-Auth-Token"] = self._token

            response = self.session.get(
                f"{self.BASE_URL}/secinfo/view/black_ip/{report_id}",
                headers=headers,
                timeout=self.TIMEOUT
            )
            response.raise_for_status()

            details = response.json()

            # Debug: Log all fields to verify correct field names
            logger.info("secudium_details_fetched",
                       report_id=report_id,
                       has_file=bool(details.get('sefileString255')),
                       all_fields=list(details.keys()),
                       file_uuid=details.get('sefileString255'),
                       filename=details.get('fileString255'))

            return details

        except requests.RequestException as e:
            logger.error("secudium_detail_failed",
                        report_id=report_id,
                        error=str(e))
            raise

    def download_excel_file_browser(self, report_title: str) -> Optional[bytes]:
        """
        Download Excel file using browser automation (Playwright).

        The API endpoint /file/SECINFO/download returns "관리자에게 문의해 주세요"
        error even with valid authentication. Browser automation is used as workaround.

        Args:
            report_title: Report title to identify which download button to click

        Returns:
            Binary Excel file content, or None if download fails

        Raises:
            RuntimeError: If not authenticated
        """
        if not self._authenticated:
            raise RuntimeError("Must authenticate before downloading files")

        try:
            from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

            logger.info("secudium_browser_download_start",
                       report_title=report_title,
                       username=self.username)

            # Create temp directory for downloads
            download_dir = tempfile.mkdtemp(prefix="secudium_")

            with sync_playwright() as p:
                # Launch headless browser
                browser = p.chromium.launch(
                    headless=True,
                    args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage']
                )

                context = browser.new_context(
                    viewport={'width': 1920, 'height': 1080},
                    accept_downloads=True
                )

                page = context.new_page()

                # Step 1: Navigate and login
                logger.info("secudium_browser_navigate")
                page.goto('https://secudium.skinfosec.co.kr', wait_until='networkidle')

                page.fill('input[name="login_name"]', self.username)
                page.fill('input[name="password"]', self.password)
                page.click('button[type="submit"]')

                # Handle duplicate session dialog
                try:
                    page.wait_for_selector('text=중복', timeout=3000)
                    logger.info("secudium_browser_duplicate_session")
                    page.click('button:has-text("예")')
                except PlaywrightTimeout:
                    pass

                page.wait_for_load_state('networkidle')
                logger.info("secudium_browser_logged_in")

                # Step 2: Navigate to BlackIP
                page.click('text=보안정보')
                page.wait_for_timeout(1000)
                page.click('text=BlackIP')
                page.wait_for_load_state('networkidle')
                logger.info("secudium_browser_navigated_blackip")

                # Step 3: Download file
                logger.info("secudium_browser_downloading",
                          report_title=report_title)

                with page.expect_download(timeout=30000) as download_info:
                    # Click the first "Down" button
                    page.locator('button:has-text("Down")').first.click()

                download = download_info.value

                # Save to temp file and read content
                temp_file = os.path.join(download_dir, download.suggested_filename)
                download.save_as(temp_file)

                with open(temp_file, 'rb') as f:
                    content = f.read()

                # Cleanup
                os.remove(temp_file)
                shutil.rmtree(download_dir, ignore_errors=True)
                browser.close()

                logger.info("secudium_browser_download_complete",
                          filename=download.suggested_filename,
                          size_bytes=len(content),
                          size_kb=len(content) / 1024)

                return content

        except Exception as e:
            logger.error("secudium_browser_download_failed",
                        report_title=report_title,
                        error=str(e),
                        error_type=type(e).__name__)
            return None

    def collect_latest_via_browser(self, limit: int = 5) -> List[Tuple[List[Dict], Dict]]:
        """
        Complete browser-based collection workflow (login + list + download + parse).

        This method is more reliable than API-based collection because:
        1. API listing endpoint returns 401 even with valid token
        2. API download endpoint returns "관리자에게 문의해 주세요" error
        3. Browser automation works consistently

        Args:
            limit: Maximum number of reports to collect (default 5)

        Returns:
            List of tuples: [(records, metadata), (records, metadata), ...]
            where records is list of blacklist IP dictionaries
            and metadata contains report info
        """
        try:
            from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

            logger.info("secudium_browser_collection_start",
                       username=self.username,
                       limit=limit)

            results = []
            download_dir = tempfile.mkdtemp(prefix="secudium_")

            with sync_playwright() as p:
                # Launch browser
                browser = p.chromium.launch(
                    headless=True,
                    args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage']
                )

                context = browser.new_context(
                    viewport={'width': 1920, 'height': 1080},
                    accept_downloads=True
                )

                page = context.new_page()

                # Step 1: Login
                logger.info("secudium_browser_login_start")
                page.goto('https://secudium.skinfosec.co.kr', wait_until='networkidle', timeout=60000)

                # Debug: Take screenshot and log page content
                page.screenshot(path=f'{download_dir}/login_page.png')
                page_html = page.content()
                logger.info("secudium_page_loaded",
                           html_length=len(page_html),
                           has_login_form='name="login_name"' in page_html,
                           has_submit_button='type="submit"' in page_html)

                # Wait for login form to be ready
                page.wait_for_selector('input[name="login_name"]', state='visible', timeout=10000)
                page.fill('input[name="login_name"]', self.username)
                page.fill('input[name="password"]', self.password)

                # Try multiple selectors for submit button
                try:
                    # Try standard button selector
                    page.wait_for_selector('button[type="submit"]', state='visible', timeout=3000)
                    page.click('button[type="submit"]', timeout=60000)
                except PlaywrightTimeout:
                    # Try alternative selectors
                    logger.info("secudium_trying_alternative_submit_selector")
                    page.screenshot(path=f'{download_dir}/before_submit.png')

                    # Try login button by text or class
                    if page.locator('button:has-text("로그인")').count() > 0:
                        page.click('button:has-text("로그인")', timeout=60000)
                    elif page.locator('input[type="submit"]').count() > 0:
                        page.click('input[type="submit"]', timeout=60000)
                    elif page.locator('.btn-login').count() > 0:
                        page.click('.btn-login', timeout=60000)
                    else:
                        # Last resort: press Enter
                        logger.info("secudium_submit_via_enter")
                        page.press('input[name="password"]', 'Enter')
                        page.wait_for_timeout(2000)

                # Handle duplicate session dialog (appears after login attempt)
                # Dialog text: "동일 ID로 로그인 한 사용자가 있습니다..."
                # Buttons: "Yes" / "No" (English text)
                try:
                    # Wait for dialog to appear (try multiple selectors)
                    dialog_found = False
                    dialog_selectors = [
                        '.dhtmlx_popup_text',
                        'text=동일',
                        '.dhtmlx_modal_box',
                        '[dhxbox]'
                    ]

                    for selector in dialog_selectors:
                        try:
                            page.wait_for_selector(selector, timeout=3000)
                            dialog_found = True
                            break
                        except PlaywrightTimeout:
                            continue

                    if not dialog_found:
                        raise PlaywrightTimeout("No duplicate session dialog found")

                    logger.info("secudium_browser_duplicate_session_dialog_found")

                    # Click Yes button (English text, not Korean)
                    yes_selectors = [
                        'button:has-text("Yes")',
                        'div.dhtmlx_popup_button:has-text("Yes")',
                        '[result="true"]',
                        'button:has-text("예")'  # Fallback to Korean
                    ]

                    yes_clicked = False
                    for selector in yes_selectors:
                        try:
                            page.click(selector, timeout=3000)
                            logger.info("secudium_duplicate_session_confirmed", selector=selector)
                            yes_clicked = True
                            break
                        except PlaywrightTimeout:
                            continue

                    # After clicking Yes, wait for actual navigation away from login page
                    if yes_clicked:
                        try:
                            # Wait for URL to change OR title to change (max 15 seconds)
                            page.wait_for_function(
                                "() => !document.title.includes('Login') || window.location.href.includes('main') || window.location.href.includes('dashboard')",
                                timeout=15000
                            )
                            logger.info("secudium_navigation_after_dialog")
                        except PlaywrightTimeout:
                            logger.warning("secudium_dialog_navigation_timeout",
                                          url=page.url,
                                          title=page.title())
                            # Continue anyway, might still work
                            pass

                    page.wait_for_timeout(2000)
                except PlaywrightTimeout:
                    logger.info("secudium_no_duplicate_session_dialog")
                    pass

                page.wait_for_load_state('networkidle')

                # Wait for dashboard to fully load after login
                page.wait_for_timeout(3000)

                # Verify we're actually logged in (not still on login page)
                current_url = page.url
                page_title = page.title()
                is_still_login_page = 'Login' in page_title or page.locator('input[name="login_name"]').count() > 0

                if is_still_login_page:
                    logger.error("secudium_login_failed_still_on_login_page",
                                url=current_url,
                                title=page_title)
                    raise RuntimeError("Login failed: Still on login page after authentication")

                logger.info("secudium_browser_login_complete", url=current_url, title=page_title)

                # Debug: See what's on the page after login
                page.screenshot(path=f'{download_dir}/after_login.png')
                page_html = page.content()

                # Save dashboard HTML for inspection
                with open(f'{download_dir}/dashboard.html', 'w', encoding='utf-8') as f:
                    f.write(page_html)

                logger.info("secudium_after_login",
                           url=page.url,
                           html_length=len(page_html),
                           html_has_menu='보안정보' in page_html,
                           html_has_blackip='BlackIP' in page_html,
                           html_has_secinfo='secinfo' in page_html.lower(),
                           html_has_iframe='<iframe' in page_html,
                           html_snippet=page_html[:500])

                # Step 2: Navigate to BlackIP (MUST use menu, direct URL breaks session)
                logger.info("secudium_browser_navigate_blackip")

                # Strategy 1: Try menu navigation
                menu_clicked = False
                security_menu_selectors = [
                    'text=보안정보',
                    'a:has-text("보안정보")',
                    'a[href*="secinfo"]',
                    '#menu-security',
                    '.menu-security',
                    'li:has-text("보안정보")',
                    '[class*="menu"]:has-text("보안정보")'
                ]

                for selector in security_menu_selectors:
                    try:
                        page.wait_for_selector(selector, state='visible', timeout=5000)
                        page.click(selector, timeout=10000)
                        logger.info("secudium_menu_clicked", selector=selector)
                        menu_clicked = True
                        break
                    except PlaywrightTimeout:
                        continue

                if menu_clicked:
                    page.wait_for_timeout(2000)

                    # Click BlackIP submenu
                    blackip_selectors = [
                        'text=BlackIP',
                        'a:has-text("BlackIP")',
                        'a[href*="black_ip"]',
                        'li:has-text("BlackIP")'
                    ]

                    for selector in blackip_selectors:
                        try:
                            page.wait_for_selector(selector, state='visible', timeout=5000)
                            page.click(selector, timeout=10000)
                            logger.info("secudium_blackip_clicked", selector=selector)
                            break
                        except PlaywrightTimeout:
                            continue

                    page.wait_for_load_state('networkidle', timeout=60000)

                # Strategy 2: If menu navigation failed, try JavaScript navigation
                if not menu_clicked:
                    logger.info("secudium_menu_nav_failed_trying_js")
                    try:
                        # Try navigating via JavaScript (preserves session)
                        page.evaluate("""
                            () => {
                                // Try finding and clicking menu items via JavaScript
                                const menuItems = document.querySelectorAll('a, li, div, span');
                                for (const item of menuItems) {
                                    if (item.textContent.includes('보안정보') || item.textContent.includes('BlackIP')) {
                                        item.click();
                                        return true;
                                    }
                                }
                                return false;
                            }
                        """)
                        page.wait_for_timeout(3000)
                        logger.info("secudium_js_nav_attempted")
                    except Exception as e:
                        logger.error("secudium_js_nav_failed", error=str(e))

                # Verify we're on the BlackIP page
                current_url = page.url
                current_html = page.content()
                is_blackip_page = 'black_ip' in current_url.lower() or 'blackip' in current_html.lower()

                logger.info("secudium_nav_verification",
                           url=current_url,
                           is_blackip_page=is_blackip_page,
                           html_length=len(current_html))

                if not is_blackip_page:
                    logger.error("secudium_nav_failed_not_on_blackip_page", url=current_url)
                    # Return empty result instead of failing
                    return []

                # Step 3: Download files (up to limit)
                logger.info("secudium_browser_download_files", limit=limit)

                # Debug: Take screenshot of BlackIP page
                page.screenshot(path=f'{download_dir}/blackip_page.png')
                page_html = page.content()

                # Save HTML for inspection
                html_file = f'{download_dir}/blackip_page.html'
                with open(html_file, 'w', encoding='utf-8') as f:
                    f.write(page_html)

                logger.info("secudium_blackip_page_loaded",
                           url=page.url,
                           html_length=len(page_html),
                           has_download='download' in page_html.lower(),
                           has_다운로드='다운로드' in page_html,
                           html_snippet=page_html[:800],  # First 800 chars
                           html_file=html_file)

                # Wait for dynamic content to load
                page.wait_for_timeout(2000)

                # Get download buttons (try multiple strategies)
                download_buttons = []

                # Strategy 1: Look for download buttons/links (REDUCED from 9 to 3 selectors to prevent IP blocking)
                selectors_to_try = [
                    'button:has-text("다운로드")',  # Primary: Korean download button
                    'a:has-text("다운로드")',      # Secondary: Korean download link
                    '[class*="download"]'          # Fallback: Download class
                ]

                logger.info("secudium_selector_search_start",
                           selectors=len(selectors_to_try),
                           note="Reduced from 9 to 3 to prevent excessive requests")

                for selector in selectors_to_try:
                    try:
                        buttons = page.locator(selector).all()
                        if buttons:
                            logger.info("secudium_selector_matched", selector=selector, count=len(buttons))
                            download_buttons = buttons
                            break
                    except Exception as e:
                        logger.debug("secudium_selector_failed", selector=selector, error=str(e))
                        continue

                # Strategy 2: If no buttons found, try table rows (common pattern)
                if not download_buttons:
                    try:
                        table_rows = page.locator('table tbody tr').all()
                        logger.info("secudium_table_rows_found", count=len(table_rows))
                        if table_rows:
                            download_buttons = table_rows  # Use rows as clickable elements
                    except Exception as e:
                        pass

                logger.info("secudium_download_elements_found", count=len(download_buttons))
                actual_limit = min(limit, len(download_buttons))

                # Anti-blocking measures: Track consecutive failures
                consecutive_failures = 0
                MAX_CONSECUTIVE_FAILURES = 3

                for i in range(actual_limit):
                    try:
                        # Check consecutive failure limit (prevent IP blocking)
                        if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                            logger.warning("secudium_max_failures_reached",
                                         failures=consecutive_failures,
                                         action="stopping_collection")
                            break

                        # Re-query buttons (page may have changed)
                        buttons = page.locator('button:has-text("Down")').all()

                        # Download file
                        with page.expect_download(timeout=30000) as download_info:
                            buttons[i].click()

                        download = download_info.value
                        temp_file = os.path.join(download_dir, download.suggested_filename)
                        download.save_as(temp_file)

                        # Parse file
                        with open(temp_file, 'rb') as f:
                            content = f.read()

                        records = self.parse_excel_data(content, download.suggested_filename)
                        metadata = {
                            'report_id': None,  # Not available in browser mode
                            'title': f"Browser download {i+1}",
                            'filename': download.suggested_filename,
                            'records_count': len(records),
                            'processed_at': datetime.now(),
                            'download_method': 'browser',
                            'index': i + 1
                        }

                        results.append((records, metadata))

                        # Reset failure counter on success
                        consecutive_failures = 0

                        logger.info("secudium_browser_file_collected",
                                   filename=download.suggested_filename,
                                   records=len(records),
                                   index=i+1,
                                   total=actual_limit)

                        # Cleanup temp file
                        os.remove(temp_file)

                        # Wait before next download (INCREASED from 1s to 3s to prevent rate limiting)
                        if i < actual_limit - 1:
                            page.wait_for_timeout(3000)

                    except Exception as e:
                        consecutive_failures += 1
                        backoff_time = min(3000 * consecutive_failures, 10000)  # 3s, 6s, 9s, max 10s

                        logger.error("secudium_browser_file_failed",
                                    index=i+1,
                                    error=str(e),
                                    consecutive_failures=consecutive_failures,
                                    backoff_ms=backoff_time)

                        # Exponential backoff on failures
                        if i < actual_limit - 1:
                            logger.info("secudium_backoff_wait",
                                       duration_ms=backoff_time,
                                       reason="prevent_ip_blocking")
                            page.wait_for_timeout(backoff_time)

                        continue

                # Cleanup
                browser.close()
                shutil.rmtree(download_dir, ignore_errors=True)

                logger.info("secudium_browser_collection_complete",
                           files_collected=len(results),
                           total_records=sum(len(r[0]) for r in results))

                return results

        except Exception as e:
            logger.error("secudium_browser_collection_failed",
                        error=str(e),
                        error_type=type(e).__name__)
            import traceback
            traceback.print_exc()
            return []

    def download_excel_file(self, file_uuid: str, filename: str) -> bytes:
        """
        Download Excel file from SECUDIUM.

        NOTE: This method currently fails due to API endpoint blocking downloads
        with "관리자에게 문의해 주세요" error. Use download_excel_file_browser() instead.

        Args:
            file_uuid: File UUID from report details (sefileString255)
            filename: Original filename from report details (fileString255)

        Returns:
            Binary Excel file content

        Raises:
            RuntimeError: If not authenticated or download fails
        """
        if not self._authenticated:
            raise RuntimeError("Must authenticate before downloading files")

        logger.warning("secudium_api_download_deprecated",
                      message="API download endpoint blocked, use browser automation",
                      file_uuid=file_uuid)

        # Try API download (will likely fail)
        try:
            headers = {"X-Auth-Token": self._token} if self._token else {}
            params = {"sefileString255": file_uuid, "fileName": filename}

            response = self.session.get(
                f"{self.BASE_URL}/file/SECINFO/download",
                params=params,
                headers=headers,
                timeout=60
            )
            response.raise_for_status()

            content = response.content

            # Check for error response
            if len(content) < 1000:
                logger.error("secudium_api_download_blocked",
                           size=len(content),
                           error_message=content[:200].decode('utf-8', errors='ignore'))
                raise RuntimeError("API download blocked: 관리자에게 문의해 주세요")

            return content

        except Exception as e:
            logger.error("secudium_download_failed",
                        file_uuid=file_uuid,
                        error=str(e))
            raise

    def parse_excel_data(self, excel_content: bytes, filename: str) -> List[Dict]:
        """
        Parse Excel file and extract blacklist IP data.

        Attempts multiple parsing strategies:
        1. pandas with openpyxl engine (preferred)
        2. pandas with xlrd engine (for older .xls files)
        3. openpyxl direct parsing (fallback for corrupted files)

        Args:
            excel_content: Binary Excel file content
            filename: Original filename (for format detection)

        Returns:
            List of parsed IP records with fields:
                - ip_address: IP address string
                - country: ISO country code
                - reason: Detection reason
                - detection_date: Date IP was detected
                - removal_date: Date IP should be removed
                - data_source: Always 'SECUDIUM'
                - raw_data: Original row data as dict

        Raises:
            ValueError: If file cannot be parsed
        """
        try:
            # Validate input
            if not excel_content:
                raise ValueError("Empty Excel content received")

            if len(excel_content) < 100:  # Minimum valid Excel file size
                raise ValueError(f"Excel content too small ({len(excel_content)} bytes), likely corrupted")

            logger.info("secudium_parse_start",
                       filename=filename,
                       size=len(excel_content),
                       format=filename.split('.')[-1])

            # Try pandas first (preferred method)
            try:
                # Detect file format
                file_ext = filename.split('.')[-1].lower()
                if file_ext not in ['xls', 'xlsx']:
                    raise ValueError(f"Unsupported file format: {file_ext} (expected .xls or .xlsx)")

                engine = 'openpyxl' if file_ext == 'xlsx' else 'xlrd'

                df = pd.read_excel(
                    io.BytesIO(excel_content),
                    engine=engine,
                    header=0  # First row is header
                )

                # Validate DataFrame
                if df.empty:
                    logger.warning("secudium_empty_dataframe", filename=filename)
                    return []

                records = self._parse_dataframe(df)

                logger.info("secudium_parse_success",
                           method="pandas",
                           engine=engine,
                           rows=len(df),
                           records=len(records))

                return records

            except Exception as pandas_error:
                logger.warning("secudium_pandas_failed",
                              filename=filename,
                              error=str(pandas_error),
                              error_type=type(pandas_error).__name__,
                              fallback="openpyxl")

                # Fallback to openpyxl direct parsing
                return self._parse_excel_fallback(excel_content)

        except Exception as e:
            logger.error("secudium_parse_failed",
                        filename=filename,
                        error=str(e),
                        error_type=type(e).__name__,
                        size=len(excel_content) if excel_content else 0)
            raise ValueError(f"Failed to parse Excel file '{filename}': {str(e)}")

    def _parse_dataframe(self, df: pd.DataFrame) -> List[Dict]:
        """
        Parse pandas DataFrame into standardized record format.

        Handles multiple column name variations (Korean/English).

        Args:
            df: pandas DataFrame from Excel file

        Returns:
            List of standardized record dictionaries
        """
        records = []

        # Column name variations (case-insensitive)
        column_mappings = {
            'ip': ['ip', 'ip address', 'ip_address', 'ip주소', '아이피'],
            'country': ['country', '국가', 'country code'],
            'reason': ['reason', '사유', '탐지사유', 'detection reason', '설명'],
            'detection_date': ['detection date', '탐지일', '탐지일자', 'detect date'],
            'removal_date': ['removal date', '해제일', '해제일자', 'remove date']
        }

        # Normalize column names
        df.columns = [str(col).strip().lower() for col in df.columns]

        # Find actual column names
        actual_columns = {}
        for field, variations in column_mappings.items():
            for col in df.columns:
                if any(var.lower() in col for var in variations):
                    actual_columns[field] = col
                    break

        logger.debug("secudium_columns_mapped",
                    actual_columns=actual_columns)

        # Parse rows
        for idx, row in df.iterrows():
            try:
                # Extract IP address (required field)
                ip_address = None
                if 'ip' in actual_columns:
                    ip_address = str(row[actual_columns['ip']]).strip()

                # Skip rows without valid IP
                if not ip_address or ip_address == 'nan' or ip_address == '':
                    continue

                # Create raw_data first for fallback
                raw_data = row.to_dict()

                record = {
                    'ip_address': ip_address,
                    'country': self._safe_get(row, actual_columns, 'country'),
                    'reason': self._safe_get(row, actual_columns, 'reason'),
                    'detection_date': self._parse_date(
                        self._safe_get(row, actual_columns, 'detection_date')
                    ),
                    'removal_date': self._parse_date(
                        self._safe_get(row, actual_columns, 'removal_date')
                    ),
                    'data_source': 'SECUDIUM',
                    'raw_data': raw_data
                }

                # v3.3.5: Fallback - Extract from raw_data if column mapping failed
                if not record['detection_date'] and '탐지시간' in raw_data:
                    record['detection_date'] = self._parse_date(raw_data['탐지시간'])

                if not record['reason'] and '공격종류' in raw_data:
                    record['reason'] = str(raw_data['공격종류'])

                # Combine reason from multiple fields
                if '공격유형' in raw_data and record['reason']:
                    attack_type = str(raw_data['공격유형'])
                    if attack_type not in record['reason']:
                        record['reason'] = f"{attack_type} - {record['reason']}"

                # v3.3.5: Auto-calculate removal_date (detection_date + 3 months)
                if record['detection_date'] and not record['removal_date']:
                    from dateutil.relativedelta import relativedelta
                    record['removal_date'] = record['detection_date'] + relativedelta(months=3)

                records.append(record)

            except Exception as e:
                logger.warning("secudium_row_parse_error",
                              row_index=idx,
                              error=str(e))
                continue

        return records

    def _safe_get(self, row, column_map: Dict, field: str) -> Optional[str]:
        """Safely extract field from row using column mapping"""
        if field in column_map:
            value = row[column_map[field]]
            if pd.notna(value):
                return str(value).strip()
        return None

    def _parse_date(self, date_value) -> Optional[datetime]:
        """
        Parse date from various formats.

        Handles:
        - datetime objects
        - ISO format strings (YYYY-MM-DD)
        - Korean format strings (YYYY/MM/DD, YYYY.MM.DD)
        - Excel serial numbers
        """
        if date_value is None:
            return None

        if pd.isna(date_value):
            return None

        if isinstance(date_value, datetime):
            return date_value

        # Try common date formats (most specific first)
        date_formats = [
            '%Y-%m-%d %H:%M:%S.%f',  # v3.3.5: SECUDIUM format with fractional seconds
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y.%m.%d',
            '%Y년 %m월 %d일'
        ]

        date_str = str(date_value).strip()

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        # If all parsing fails, log warning and return None
        logger.warning("secudium_date_parse_failed", date_value=date_value)
        return None

    def _parse_excel_fallback(self, excel_content: bytes) -> List[Dict]:
        """
        Fallback parser for corrupted Excel files using openpyxl.

        This method attempts to parse files that pandas cannot handle,
        such as files with formatting issues or non-standard structures.

        Args:
            excel_content: Binary Excel file content

        Returns:
            List of parsed IP records
        """
        try:
            logger.info("secudium_fallback_parse")

            workbook = load_workbook(
                filename=io.BytesIO(excel_content),
                read_only=True,
                data_only=True  # Get cell values, not formulas
            )

            sheet = workbook.active
            records = []

            # Assume first row is header
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            # Parse data rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):  # Skip empty rows
                    continue

                row_dict = dict(zip(headers, row))

                # Extract IP (try multiple column names)
                ip_address = None
                for ip_col in ['IP', 'IP Address', 'IP주소', 'ip']:
                    if ip_col in row_dict and row_dict[ip_col]:
                        ip_address = str(row_dict[ip_col]).strip()
                        break

                if not ip_address:
                    continue

                record = {
                    'ip_address': ip_address,
                    'country': row_dict.get('Country') or row_dict.get('국가'),
                    'reason': row_dict.get('Reason') or row_dict.get('사유'),
                    'detection_date': self._parse_date(
                        row_dict.get('Detection Date') or row_dict.get('탐지일')
                    ),
                    'removal_date': self._parse_date(
                        row_dict.get('Removal Date') or row_dict.get('해제일')
                    ),
                    'data_source': 'SECUDIUM',
                    'raw_data': row_dict
                }

                records.append(record)

            logger.info("secudium_fallback_success", records=len(records))
            return records

        except Exception as e:
            logger.error("secudium_fallback_failed", error=str(e))
            raise

    def process_report(self, report_id: int, use_browser: bool = True) -> Tuple[List[Dict], Dict]:
        """
        Complete processing of a SECUDIUM report.

        This is a convenience method that:
        1. Gets report details
        2. Downloads Excel file (via browser automation by default)
        3. Parses data
        4. Returns records and metadata

        Args:
            report_id: SECUDIUM report ID
            use_browser: Use browser automation for download (default: True)
                        API download endpoint is blocked, browser is recommended

        Returns:
            Tuple of (records, metadata):
                - records: List of parsed IP records
                - metadata: Report metadata dictionary
        """
        logger.info("secudium_process_report",
                   report_id=report_id,
                   use_browser=use_browser)

        # Get report details
        details = self.get_report_details(report_id)

        # Extract file info
        file_uuid = details.get('sefileString255')
        filename = details.get('fileString255')
        report_title = details.get('title', '')

        if not file_uuid or not filename:
            raise ValueError(f"Report {report_id} has no downloadable file")

        # Download Excel file with fallback strategy
        excel_content = None

        if use_browser:
            # Primary: Browser automation (works reliably)
            logger.info("secudium_trying_browser_download")
            excel_content = self.download_excel_file_browser(report_title)

            if not excel_content or len(excel_content) < 1000:
                logger.warning("secudium_browser_failed_trying_api")
                # Fallback to API (likely to fail but worth trying)
                try:
                    excel_content = self.download_excel_file(file_uuid, filename)
                except Exception as api_error:
                    logger.error("secudium_api_download_also_failed",
                               error=str(api_error))
                    raise ValueError("Both browser and API downloads failed")
        else:
            # Try API first, fallback to browser
            logger.info("secudium_trying_api_download")
            try:
                excel_content = self.download_excel_file(file_uuid, filename)
            except Exception as api_error:
                logger.warning("secudium_api_failed_trying_browser",
                             error=str(api_error))
                excel_content = self.download_excel_file_browser(report_title)

        if not excel_content or len(excel_content) < 1000:
            raise ValueError("Failed to download Excel file from SECUDIUM")

        # Parse data
        records = self.parse_excel_data(excel_content, filename)

        # Metadata
        metadata = {
            'report_id': report_id,
            'title': details.get('title'),
            'filename': filename,
            'records_count': len(records),
            'processed_at': datetime.now(),
            'download_method': 'browser' if use_browser else 'api'
        }

        logger.info("secudium_process_complete",
                   report_id=report_id,
                   records=len(records),
                   method=metadata['download_method'])

        return records, metadata

    def close(self):
        """Close session and cleanup resources"""
        self.session.close()
        logger.info("secudium_client_closed")


def collect_secudium_data(db_service, credentials: Dict) -> Dict:
    """
    Main collection function for SECUDIUM data.

    This function orchestrates the complete collection process:
    1. Authenticate with SECUDIUM
    2. Fetch latest reports
    3. Process each report (download + parse)
    4. Store in database
    5. Record collection history

    Args:
        db_service: Database service instance
        credentials: SECUDIUM credentials dict with 'username' and 'password'

    Returns:
        Collection statistics dictionary:
            - source: 'SECUDIUM'
            - reports_processed: Number of reports processed
            - ips_collected: Total IPs collected
            - ips_new: New IPs added
            - ips_duplicate: Duplicate IPs skipped
            - timestamp: Collection timestamp
            - errors: List of errors encountered

    Example:
        >>> from core.services.database_service import service as db_service
        >>> credentials = {'username': 'user', 'password': 'pass'}
        >>> result = collect_secudium_data(db_service, credentials)
        >>> print(f"Collected {result['ips_collected']} IPs from {result['reports_processed']} reports")
    """
    client = None

    try:
        logger.info("secudium_collection_start")

        # Initialize client
        client = SecudiumAPIClient(
            username=credentials['username'],
            password=credentials['password']
        )

        # Authenticate
        client.authenticate()

        # Get latest reports (last 5 days to avoid re-processing)
        reports = client.get_latest_reports(limit=5, days_back=5)

        if not reports:
            logger.info("secudium_no_reports")
            return {
                'source': 'SECUDIUM',
                'reports_processed': 0,
                'ips_collected': 0,
                'timestamp': datetime.now()
            }

        total_ips = 0
        new_ips = 0
        duplicate_ips = 0
        errors = []

        for report in reports:
            try:
                report_id = report['id']

                # Check if already processed
                if db_service.is_report_processed('SECUDIUM', report_id):
                    logger.info("secudium_report_skipped",
                               report_id=report_id,
                               reason="already_processed")
                    continue

                # Process report
                records, metadata = client.process_report(report_id)

                # Insert records into database
                result = db_service.batch_insert_blacklist_ips(records)

                new_ips += result.get('inserted', 0)
                duplicate_ips += result.get('duplicates', 0)
                total_ips += len(records)

                # Mark report as processed
                db_service.mark_report_processed('SECUDIUM', report_id, metadata)

                logger.info("secudium_report_processed",
                           report_id=report_id,
                           ips=len(records),
                           new=result.get('inserted', 0))

            except Exception as e:
                error_msg = f"Report {report['id']}: {str(e)}"
                errors.append(error_msg)
                logger.error("secudium_report_error",
                            report_id=report['id'],
                            error=str(e))
                continue

        # Collection summary
        result = {
            'source': 'SECUDIUM',
            'reports_processed': len([r for r in reports if not db_service.is_report_processed('SECUDIUM', r['id'])]),
            'ips_collected': total_ips,
            'ips_new': new_ips,
            'ips_duplicate': duplicate_ips,
            'timestamp': datetime.now(),
            'errors': errors
        }

        logger.info("secudium_collection_complete",
                   reports=result['reports_processed'],
                   ips_total=total_ips,
                   ips_new=new_ips,
                   ips_duplicate=duplicate_ips)

        return result

    except Exception as e:
        logger.error("secudium_collection_failed", error=str(e))
        raise

    finally:
        if client:
            client.close()


if __name__ == "__main__":
    """
    Manual testing script.

    Usage:
        python -m collector.api.secudium_api
    """
    import os
    import sys

    # Load credentials from environment
    username = os.getenv('SECUDIUM_USERNAME')
    password = os.getenv('SECUDIUM_PASSWORD')

    if not username or not password:
        print("Error: SECUDIUM_USERNAME and SECUDIUM_PASSWORD must be set")
        sys.exit(1)

    # Test collection
    try:
        print("Testing SECUDIUM collection...")

        # Mock database service for testing
        class MockDBService:
            def is_report_processed(self, source, report_id):
                return False

            def batch_insert_blacklist_ips(self, records):
                print(f"Would insert {len(records)} records")
                return {'inserted': len(records), 'duplicates': 0}

            def mark_report_processed(self, source, report_id, metadata):
                print(f"Marked report {report_id} as processed")

        db_service = MockDBService()
        credentials = {'username': username, 'password': password}

        result = collect_secudium_data(db_service, credentials)

        print("\n" + "="*50)
        print("SECUDIUM Collection Result:")
        print("="*50)
        print(f"Reports processed: {result['reports_processed']}")
        print(f"IPs collected: {result['ips_collected']}")
        print(f"New IPs: {result['ips_new']}")
        print(f"Duplicate IPs: {result['ips_duplicate']}")
        print(f"Timestamp: {result['timestamp']}")

        if result['errors']:
            print(f"\nErrors encountered: {len(result['errors'])}")
            for error in result['errors']:
                print(f"  - {error}")

        print("="*50)

    except Exception as e:
        print(f"\nCollection failed: {str(e)}")
        sys.exit(1)
